from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import random
from openai import RateLimitError
import time
import json
import pandas as pd
import math
import fitment_logic

def pixels_to_excel_width(px):
    return px / 7

def format_products_workbook(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    column_widths_px = {
        "Title": 260,
        "Tags": 260,
        "Metafield: convermax.fitment [list.single_line_text_field]": 360,
        "AI Review Reasons": 260,
        "Custom Collections": 160,
        "Variant SKU": 170,
        "Type": 120,
    }

    review_fill = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    for cell in ws[1]:
        col_name = cell.value
        if col_name in column_widths_px:
            col_letter = get_column_letter(cell.column)
            ws.column_dimensions[col_letter].width = pixels_to_excel_width(
                column_widths_px[col_name]
            )

            for row_cell in ws[col_letter]:
                row_cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

            if col_name == "AI Review Reasons":
                for row_cell in ws[col_letter]:
                    row_cell.fill = review_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)

def safe_call_ai_for_row(client, prompt_id, prompt_version, row_index, row, max_retries=6):
    for attempt in range(max_retries):
        try:
            return call_ai_for_row(
                client,
                prompt_id,
                prompt_version,
                row_index,
                row
            )

        except RateLimitError as e:
            wait = min(60, (2 ** attempt) + random.uniform(0, 1))
            print(e)
            time.sleep(wait)

    raise RuntimeError(f"Failed row {row_index} after {max_retries} retries")

def format_tags(tags):
    """
    Converts a list of tags back into a clean Shopify comma-separated tag string.

    Also:
    - trims whitespace
    - removes empty tags
    - removes duplicate tags
    - sorts tags for stable output
    """

    tags = [str(tag).strip() for tag in tags if str(tag).strip()]
    return ", ".join(sorted(set(tags)))

def call_ai_for_row(client, prompt_id, prompt_version, row_index, row):
    """
    Sends one spreadsheet row to OpenAI and returns parsed JSON.
    Returns:
        {
            "row_index": row_index,
            "ai_result": parsed_ai_json
        }
    """

    product = build_product_input(row)

    response = client.responses.create(
        prompt={
            "id": prompt_id,
            "version": prompt_version,
            "variables": {
                "product": json.dumps(product, ensure_ascii=False)
            }
        }
    )

    return {
        "row_index": row_index,
        "ai_result": json.loads(response.output_text)
    }

def parse_tags(value):
    """
    Converts Shopify comma-separated tags into a Python list.
    Example: "BMW, Exhaust, M340i"
    Returns: ["BMW", "Exhaust", "M340i"]
    """

    if pd.isna(value) or str(value).strip() == "":
        return []

    return [
        tag.strip()
        for tag in str(value).split(",")
        if tag.strip()
    ]

def build_product_input(row):
    """
    Converts a spreadsheet row into the JSON object that is sent to OpenAI.
    This is the ONLY place where spreadsheet columns are mapped into AI input fields.
    """

    return {
        "title": "" if pd.isna(row["Title"]) else str(row["Title"]),
        "body_html": "" if pd.isna(row["Body HTML"]) else str(row["Body HTML"]),
        "vendor": "" if pd.isna(row["Vendor"]) else str(row["Vendor"]),
        "tags": parse_tags(row["Tags"]),
        "current_weight": (
            0
            if pd.isna(row["Variant Weight"])
            else float(row["Variant Weight"])
        ),
        "price": 0 if pd.isna(row["Variant Price"]) else float(row["Variant Price"]),
    }

def build_review_notes(
    ai,
    fitment_review_notes,
    final_metadata,
):
    """
    Combines AI review reasons and fitment review reasons,
    adds internal consistency checks, and filters ignored reasons.

    Returns:
        review_notes,
        ignored_review_notes
    """

    review_notes = []

    if ai.get("review_required"):
        review_notes.extend(ai.get("review_reasons", []))

    review_notes.extend(fitment_review_notes)

    # This should never happen.
    if (
        ai.get("convermax_universal") is not True
        and ai.get("vehicle_fitments")
        and not final_metadata
        and not fitment_review_notes
    ):
        review_notes.append(
            "fitment_metadata_empty_for_vehicle_specific_product"
        )

    return filter_review_notes(review_notes)

def filter_review_notes(review_notes):
    """
    Removes ignored review notes.
    Returns:
        filtered_review_notes,
        ignored_review_notes
    """

    ignored_review_reasons = {
        "title_price_description_scope_conflict",
        "vehicle_fitment_submodel_unclear",
        "vehicle_fitment_submodel_ambiguous",
        "weight_low_confidence",
        "fitment_year_conflict",
    }

    ignored_review_reason_prefixes = (
        "fitment_conflict",
    )

    filtered = []
    ignored = []

    for note in sorted(set(review_notes)):
        if (
            note in ignored_review_reasons
            or any(note.startswith(prefix) for prefix in ignored_review_reason_prefixes)
        ):
            ignored.append(note)
        else:
            filtered.append(note)

    return filtered, ignored

def print_product_summary(
    *,
    debug,
    row_index,
    original_row,
    ai,
    final_metadata,
    fitment_tag,
    review_notes,
    ignored_review_notes,
    created_custom_fitments,
    title_col,
    type_col,
    weight_col,
):
    """
    Prints processing information for one product.

    When debug=True:
        Prints the full product summary.

    When debug=False:
        Prints products that created custom fitments
        or produced a no-fitment review reason.
    """

    if debug:
        print("\n" + "=" * 100)
        print(f"ROW: {row_index}")
        print(f"ID: {original_row['ID']}")
        print(f"SKU: {original_row['Variant SKU']}")
        print(f"OLD TITLE: {original_row[title_col]}")
        print(f"NEW TITLE: {ai.get('title', original_row[title_col])}")
        print(f"TYPE: {ai.get('type', original_row[type_col])}")
        print(f"COLLECTIONS: {', '.join(ai.get('custom_collections', []))}")
        print(f"WEIGHT: {ai.get('variant_weight', original_row[weight_col])}")
        print(f"METADATA: {final_metadata}")
        print(f"FITMENT TAG: {fitment_tag}")

        if ai.get("review_required"):
            print(f"AI REVIEW REASONS: {ai.get('review_reasons', [])}")

        if review_notes:
            print(f"REVIEW REASONS: {review_notes}")

        if ignored_review_notes:
            print(f"IGNORED REVIEW REASONS: {ignored_review_notes}")

    else:
        no_fitment_review_notes = [
            note
            for note in review_notes
            if note.startswith("No fitment found:")
        ]

        if created_custom_fitments or no_fitment_review_notes:
            print("\n" + "=" * 100)

            if created_custom_fitments:
                print(f"CUSTOM FITMENT CREATED FOR ROW: {row_index}")

            if no_fitment_review_notes:
                print(f"NO FITMENT FOUND FOR ROW: {row_index}")

            print(f"SKU: {original_row['Variant SKU']}")
            print("AI RESPONSE:")
            print(json.dumps(ai, indent=2, ensure_ascii=False))

            if no_fitment_review_notes:
                print(f"REVIEW REASONS: {no_fitment_review_notes}")

            print(f"METADATA: {final_metadata}")


def update_product_row(
    *,
    df,
    row_index,
    original_row,
    ai,
    review_notes,
    fitment_tag,
    final_metadata,
    review_col,
    command_col,
    tags_command_col,
    variant_command_col,
    title_col,
    type_col,
    tags_col,
    collections_col,
    weight_col,
    fitment_col,
    universal_col,
):
    """
    Updates one product row with AI enrichment and fitment results.
    """

    existing_tags = [
        tag
        for tag in parse_tags(original_row[tags_col])
        if not tag.startswith("fits_")
    ]

    if fitment_tag:
        existing_tags.append(fitment_tag)

    df.at[row_index, review_col] = (" | ".join(review_notes) if review_notes else None)
    df.at[row_index, command_col] = "MERGE"
    df.at[row_index, tags_command_col] = "REPLACE"
    df.at[row_index, variant_command_col] = "MERGE"
    df.at[row_index, title_col] = ai.get("title", original_row[title_col])
    df.at[row_index, type_col] = ai.get("type", original_row[type_col])
    df.at[row_index, tags_col] = format_tags(existing_tags)
    df.at[row_index, collections_col] = ", ".join(ai.get("custom_collections", []))
    
    weight = ai.get("variant_weight", original_row[weight_col])
    if pd.notna(weight):
        weight = math.ceil(float(weight))
    df.at[row_index, weight_col] = weight

    df.at[row_index, fitment_col] = json.dumps(final_metadata, ensure_ascii=False)
    df.at[row_index, universal_col] = ("Yes" if ai.get("convermax_universal") is True else None)


def manually_resolve_fitments(
    *,
    df,
    fitment_df,
    custom_fitment_df,
    custom_fitment_file,
    row_index,
    manual_fitments,
    ai_result,
    review_col,
    command_col,
    tags_command_col,
    variant_command_col,
    title_col,
    type_col,
    tags_col,
    collections_col,
    weight_col,
    fitment_col,
    universal_col,
    debug,
):
    """
    Manually resolves fitments for a single product.
    """
    original_row = df.loc[row_index]

    (
        final_metadata,
        custom_fitment_df,
        fitment_review_notes,
        created_custom_fitments,
    ) = fitment_logic.resolve_vehicle_fitments_to_metadata(
        fitment_df=fitment_df,
        vehicle_fitments=manual_fitments,
        custom_fitment_df=custom_fitment_df,
        custom_fitment_file=custom_fitment_file,
        product_row=original_row,
    )

    fitment_tag = fitment_logic.metadata_to_fitment_tag(final_metadata)

    review_notes, ignored_review_notes = build_review_notes(
        ai_result,
        fitment_review_notes,
        final_metadata,
    )

    update_product_row(
        df=df,
        row_index=row_index,
        original_row=original_row,
        ai=ai_result,
        review_notes=review_notes,
        fitment_tag=fitment_tag,
        final_metadata=final_metadata,
        review_col=review_col,
        command_col=command_col,
        tags_command_col=tags_command_col,
        variant_command_col=variant_command_col,
        title_col=title_col,
        type_col=type_col,
        tags_col=tags_col,
        collections_col=collections_col,
        weight_col=weight_col,
        fitment_col=fitment_col,
        universal_col=universal_col,
    )

    print_product_summary(
        debug=debug,
        row_index=row_index,
        original_row=original_row,
        ai=ai_result,
        final_metadata=final_metadata,
        fitment_tag=fitment_tag,
        review_notes=review_notes,
        ignored_review_notes=ignored_review_notes,
        created_custom_fitments=created_custom_fitments,
        title_col=title_col,
        type_col=type_col,
        weight_col=weight_col,
    )

    return (
        custom_fitment_df,
        final_metadata,
        fitment_tag,
        created_custom_fitments,
    )